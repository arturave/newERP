"""
Email Service - Modul wysylania maili
=====================================
Obsluga IMAP, SMTP i Microsoft Exchange dla wysylania:
- Faktur PDF
- Zestawien XLSX z pozycjami, grafikami, cenami
- Dokumentow WZ

Konfiguracja w config/email_config.json
"""

import os
import json
import logging
import smtplib
import imaplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from typing import List, Dict, Optional, Union
from dataclasses import dataclass, field
from datetime import datetime

logger = logging.getLogger(__name__)

# Sciezka do konfiguracji
CONFIG_PATH = Path(__file__).parent.parent / "config" / "email_config.json"


@dataclass
class EmailConfig:
    """Konfiguracja email"""
    provider: str = "smtp"  # smtp, exchange, office365
    smtp_server: str = ""
    smtp_port: int = 587
    imap_server: str = ""
    imap_port: int = 993
    username: str = ""
    password: str = ""
    use_tls: bool = True
    from_email: str = ""
    from_name: str = "NewERP System"
    # Exchange/Office365
    client_id: str = ""
    client_secret: str = ""
    tenant_id: str = ""


@dataclass
class EmailAttachment:
    """Zalacznik email"""
    filename: str
    content: bytes
    content_type: str = "application/octet-stream"


@dataclass
class EmailMessage:
    """Wiadomosc email"""
    to: List[str]
    subject: str
    body_html: str
    body_text: str = ""
    cc: List[str] = field(default_factory=list)
    bcc: List[str] = field(default_factory=list)
    attachments: List[EmailAttachment] = field(default_factory=list)
    reply_to: str = ""


class EmailService:
    """Serwis wysylania maili"""

    def __init__(self, config: EmailConfig = None):
        self.config = config or self._load_config()
        self._smtp_connection = None
        self._imap_connection = None

    def _load_config(self) -> EmailConfig:
        """Wczytaj konfiguracje z pliku"""
        if CONFIG_PATH.exists():
            try:
                with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return EmailConfig(**data)
            except Exception as e:
                logger.error(f"Error loading email config: {e}")

        # Domyslna konfiguracja
        return EmailConfig()

    def save_config(self):
        """Zapisz konfiguracje do pliku"""
        CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        try:
            with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump(self.config.__dict__, f, indent=2)
            logger.info(f"Email config saved to {CONFIG_PATH}")
        except Exception as e:
            logger.error(f"Error saving email config: {e}")

    # ========================================
    # SMTP - Wysylanie
    # ========================================

    def connect_smtp(self) -> bool:
        """Polacz z serwerem SMTP"""
        try:
            if self.config.use_tls:
                self._smtp_connection = smtplib.SMTP(
                    self.config.smtp_server,
                    self.config.smtp_port
                )
                self._smtp_connection.starttls()
            else:
                self._smtp_connection = smtplib.SMTP_SSL(
                    self.config.smtp_server,
                    self.config.smtp_port
                )

            self._smtp_connection.login(
                self.config.username,
                self.config.password
            )
            logger.info(f"Connected to SMTP: {self.config.smtp_server}")
            return True

        except Exception as e:
            logger.error(f"SMTP connection error: {e}")
            return False

    def disconnect_smtp(self):
        """Rozlacz z SMTP"""
        if self._smtp_connection:
            try:
                self._smtp_connection.quit()
            except:
                pass
            self._smtp_connection = None

    def send_email(self, message: EmailMessage) -> bool:
        """Wyslij email przez SMTP"""
        try:
            # Polacz jesli potrzeba
            if not self._smtp_connection:
                if not self.connect_smtp():
                    return False

            # Stworz wiadomosc
            msg = MIMEMultipart('alternative')
            msg['Subject'] = message.subject
            msg['From'] = f"{self.config.from_name} <{self.config.from_email}>"
            msg['To'] = ', '.join(message.to)

            if message.cc:
                msg['Cc'] = ', '.join(message.cc)
            if message.reply_to:
                msg['Reply-To'] = message.reply_to

            # Tresc
            if message.body_text:
                msg.attach(MIMEText(message.body_text, 'plain', 'utf-8'))
            msg.attach(MIMEText(message.body_html, 'html', 'utf-8'))

            # Zalaczniki
            for attachment in message.attachments:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.content)
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{attachment.filename}"'
                )
                msg.attach(part)

            # Wyslij
            all_recipients = message.to + message.cc + message.bcc
            self._smtp_connection.sendmail(
                self.config.from_email,
                all_recipients,
                msg.as_string()
            )

            logger.info(f"Email sent to: {', '.join(message.to)}")
            return True

        except Exception as e:
            logger.error(f"Error sending email: {e}")
            return False

    # ========================================
    # IMAP - Odbieranie (opcjonalne)
    # ========================================

    def connect_imap(self) -> bool:
        """Polacz z serwerem IMAP"""
        try:
            self._imap_connection = imaplib.IMAP4_SSL(
                self.config.imap_server,
                self.config.imap_port
            )
            self._imap_connection.login(
                self.config.username,
                self.config.password
            )
            logger.info(f"Connected to IMAP: {self.config.imap_server}")
            return True

        except Exception as e:
            logger.error(f"IMAP connection error: {e}")
            return False

    def disconnect_imap(self):
        """Rozlacz z IMAP"""
        if self._imap_connection:
            try:
                self._imap_connection.logout()
            except:
                pass
            self._imap_connection = None

    # ========================================
    # Microsoft Exchange / Office 365
    # ========================================

    def send_via_exchange(self, message: EmailMessage) -> bool:
        """Wyslij email przez Microsoft Exchange/Office 365"""
        try:
            # Wymaga pakietu exchangelib lub msal
            try:
                from exchangelib import (
                    Credentials, Account, Message as ExchangeMessage,
                    FileAttachment, HTMLBody, DELEGATE
                )
            except ImportError:
                logger.error("exchangelib not installed. Install with: pip install exchangelib")
                return False

            credentials = Credentials(
                username=self.config.username,
                password=self.config.password
            )

            account = Account(
                primary_smtp_address=self.config.from_email,
                credentials=credentials,
                autodiscover=True,
                access_type=DELEGATE
            )

            # Stworz wiadomosc
            msg = ExchangeMessage(
                account=account,
                subject=message.subject,
                body=HTMLBody(message.body_html),
                to_recipients=message.to
            )

            if message.cc:
                msg.cc_recipients = message.cc

            # Zalaczniki
            for attachment in message.attachments:
                file_attachment = FileAttachment(
                    name=attachment.filename,
                    content=attachment.content
                )
                msg.attach(file_attachment)

            msg.send()
            logger.info(f"Email sent via Exchange to: {', '.join(message.to)}")
            return True

        except Exception as e:
            logger.error(f"Exchange send error: {e}")
            return False

    # ========================================
    # High-level API
    # ========================================

    def send(self, message: EmailMessage) -> bool:
        """Wyslij email (automatycznie wybiera metode)"""
        if self.config.provider == "exchange":
            return self.send_via_exchange(message)
        else:
            return self.send_email(message)

    def send_invoice(self, order_id: str, to_email: str,
                     invoice_pdf: bytes, xlsx_summary: bytes = None,
                     customer_name: str = "", invoice_number: str = "",
                     wz_number: str = "", transport_cost: float = 0,
                     pallet_count: int = 0, packaging_info: str = "") -> bool:
        """
        Wyslij fakture z zestawieniem

        Args:
            order_id: ID zamowienia
            to_email: Adres email odbiorcy
            invoice_pdf: Faktura PDF jako bytes
            xlsx_summary: Zestawienie XLSX jako bytes (opcjonalne)
            customer_name: Nazwa klienta
            invoice_number: Numer faktury
            wz_number: Numer WZ
            transport_cost: Koszt transportu
            pallet_count: Liczba palet
            packaging_info: Info o opakowaniach
        """
        # Przygotuj tresc
        subject = f"Faktura {invoice_number} - {customer_name}"

        body_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2>Faktura {invoice_number}</h2>

            <p>Szanowni Panstwo,</p>

            <p>W zalaczeniu przesylamy fakture oraz zestawienie do zamowienia.</p>

            <table style="border-collapse: collapse; margin: 20px 0;">
                <tr>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Numer faktury:</strong></td>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;">{invoice_number}</td>
                </tr>
                <tr>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Numer WZ:</strong></td>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;">{wz_number}</td>
                </tr>
                <tr>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Koszt transportu:</strong></td>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;">{transport_cost:.2f} PLN</td>
                </tr>
                <tr>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Liczba palet:</strong></td>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;">{pallet_count}</td>
                </tr>
                <tr>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;"><strong>Opakowania:</strong></td>
                    <td style="padding: 5px 15px; border: 1px solid #ddd;">{packaging_info}</td>
                </tr>
            </table>

            <p>W razie pytan prosimy o kontakt.</p>

            <p>Z powazaniem,<br/>
            {self.config.from_name}</p>

            <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
            <p style="font-size: 11px; color: #888;">
                Wiadomosc wygenerowana automatycznie przez system NewERP.<br/>
                ID zamowienia: {order_id}
            </p>
        </body>
        </html>
        """

        body_text = f"""
Faktura {invoice_number}

Szanowni Panstwo,

W zalaczeniu przesylamy fakture oraz zestawienie do zamowienia.

Numer faktury: {invoice_number}
Numer WZ: {wz_number}
Koszt transportu: {transport_cost:.2f} PLN
Liczba palet: {pallet_count}
Opakowania: {packaging_info}

W razie pytan prosimy o kontakt.

Z powazaniem,
{self.config.from_name}

---
ID zamowienia: {order_id}
        """

        # Zalaczniki
        attachments = [
            EmailAttachment(
                filename=f"Faktura_{invoice_number}.pdf",
                content=invoice_pdf,
                content_type="application/pdf"
            )
        ]

        if xlsx_summary:
            attachments.append(
                EmailAttachment(
                    filename=f"Zestawienie_{invoice_number}.xlsx",
                    content=xlsx_summary,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            )

        # Wyslij
        message = EmailMessage(
            to=[to_email],
            subject=subject,
            body_html=body_html,
            body_text=body_text,
            attachments=attachments
        )

        return self.send(message)


# ========================================
# Utility functions
# ========================================

def create_xlsx_summary(order_data: Dict, parts_data: List[Dict],
                        nesting_images: List[bytes] = None) -> bytes:
    """
    Stworz zestawienie XLSX z pozycjami zamowienia

    Args:
        order_data: Dane zamowienia
        parts_data: Lista detali z cenami
        nesting_images: Grafiki nestingu (base64 lub bytes)

    Returns:
        bytes: Plik XLSX jako bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.drawing.image import Image as XLImage
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zestawienie"

        # Style
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Naglowek zamowienia
        ws['A1'] = "ZESTAWIENIE ZAMOWIENIA"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')

        ws['A3'] = "Klient:"
        ws['B3'] = order_data.get('client', order_data.get('customer_name', ''))
        ws['A4'] = "Numer zamowienia:"
        ws['B4'] = order_data.get('name', order_data.get('title', ''))
        ws['A5'] = "Data:"
        ws['B5'] = order_data.get('date_in', '')

        # Naglowki tabeli
        headers = ['Lp.', 'Nazwa', 'Material', 'Grubosc', 'Ilosc', 'Cena/szt', 'Wartosc', 'Uwagi']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Dane detali
        total = 0
        for idx, part in enumerate(parts_data, 1):
            row = 7 + idx
            qty = part.get('quantity', 1)
            unit_price = part.get('total_unit', part.get('unit_cost', 0))
            value = qty * unit_price
            total += value

            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=part.get('name', '')).border = thin_border
            ws.cell(row=row, column=3, value=part.get('material', '')).border = thin_border
            ws.cell(row=row, column=4, value=part.get('thickness', part.get('thickness_mm', 0))).border = thin_border
            ws.cell(row=row, column=5, value=qty).border = thin_border
            ws.cell(row=row, column=6, value=f"{unit_price:.2f}").border = thin_border
            ws.cell(row=row, column=7, value=f"{value:.2f}").border = thin_border
            ws.cell(row=row, column=8, value=part.get('notes', '')).border = thin_border

        # Suma
        sum_row = 8 + len(parts_data)
        ws.cell(row=sum_row, column=6, value="RAZEM:").font = Font(bold=True)
        ws.cell(row=sum_row, column=7, value=f"{total:.2f} PLN").font = Font(bold=True)

        # Szerokosci kolumn
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20

        # Grafiki nestingu (nowy arkusz)
        if nesting_images:
            ws_img = wb.create_sheet("Grafiki nestingu")
            row = 1
            for idx, img_data in enumerate(nesting_images):
                try:
                    if isinstance(img_data, str):
                        import base64
                        img_data = base64.b64decode(img_data)
                    img_stream = BytesIO(img_data)
                    img = XLImage(img_stream)
                    img.width = 400
                    img.height = 300
                    ws_img.add_image(img, f'A{row}')
                    row += 18  # Odstep miedzy grafikami
                except Exception as e:
                    logger.warning(f"Error adding image {idx}: {e}")

        # Zapisz do bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return b''
    except Exception as e:
        logger.error(f"Error creating XLSX summary: {e}")
        return b''


# ========================================
# Test
# ========================================

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)

    # Test konfiguracji
    service = EmailService()
    print(f"Config loaded: {service.config}")

    # Test XLSX
    test_order = {
        'client': 'Test Sp. z o.o.',
        'name': 'ZAM-2024-001',
        'date_in': '2024-01-15'
    }
    test_parts = [
        {'name': 'Plyta A', 'material': 'S355', 'thickness': 3.0, 'quantity': 5, 'total_unit': 45.00},
        {'name': 'Wspornik', 'material': 'DC01', 'thickness': 1.5, 'quantity': 10, 'total_unit': 12.50},
    ]

    xlsx_bytes = create_xlsx_summary(test_order, test_parts)
    print(f"XLSX created: {len(xlsx_bytes)} bytes")
