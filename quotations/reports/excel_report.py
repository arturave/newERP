"""
Excel Report Generator
======================
Generowanie szczegółowych raportów Excel z wycen i nestingu.

Wymaga: pip install openpyxl
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List

logger = logging.getLogger(__name__)

# Sprawdź dostępność openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment,
        NamedStyle, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed. Excel reports unavailable. Run: pip install openpyxl")

from . import QuotationReport, ReportGenerator, NestingReport


class ExcelReportGenerator(ReportGenerator):
    """Generator raportów Excel"""
    
    # Kolory
    COLOR_PRIMARY = 'FF8B5CF6'  # Fioletowy
    COLOR_HEADER = 'FF1A1A1A'   # Ciemny
    COLOR_LIGHT = 'FFF3F4F6'    # Jasny szary
    COLOR_SUCCESS = 'FF22C55E'  # Zielony
    COLOR_WARNING = 'FFF59E0B'  # Pomarańczowy
    
    def __init__(self, report: QuotationReport):
        super().__init__(report)
        self.wb = None
        self._styles_created = False
    
    def _setup_styles(self):
        """Utwórz style dla dokumentu"""
        if self._styles_created or not HAS_OPENPYXL:
            return
        
        # Style nagłówków
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color=self.COLOR_PRIMARY, 
                                       end_color=self.COLOR_PRIMARY, fill_type='solid')
        
        self.subheader_font = Font(name='Arial', size=10, bold=True)
        self.subheader_fill = PatternFill(start_color=self.COLOR_LIGHT,
                                          end_color=self.COLOR_LIGHT, fill_type='solid')
        
        # Styl normalny
        self.normal_font = Font(name='Arial', size=10)
        
        # Styl dla sum
        self.total_font = Font(name='Arial', size=11, bold=True, color=self.COLOR_PRIMARY[2:])
        self.total_fill = PatternFill(start_color=self.COLOR_LIGHT,
                                      end_color=self.COLOR_LIGHT, fill_type='solid')
        
        # Obramowanie
        self.thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Wyrównanie
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        
        self._styles_created = True
    
    def generate(self, output_path: str) -> bool:
        """Generuj raport Excel"""
        if not HAS_OPENPYXL:
            logger.error("openpyxl not installed")
            return False
        
        try:
            self.wb = Workbook()
            self._setup_styles()
            
            # Usuń domyślny arkusz
            default_sheet = self.wb.active
            
            # Twórz arkusze
            self._create_summary_sheet()
            self._create_parts_sheet()
            self._create_nesting_sheet()
            self._create_costs_sheet()
            
            # Usuń pusty arkusz
            if 'Sheet' in self.wb.sheetnames:
                del self.wb['Sheet']
            
            # Zapisz
            self.wb.save(output_path)
            
            logger.info(f"Excel report generated: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel generation error: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_summary_sheet(self):
        """Arkusz podsumowania"""
        ws = self.wb.create_sheet("Podsumowanie", 0)
        
        # Tytuł
        ws.merge_cells('A1:F1')
        ws['A1'] = "WYCENA - PODSUMOWANIE"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color=self.COLOR_PRIMARY[2:])
        ws['A1'].alignment = self.center_align
        
        # Metadane
        row = 3
        metadata = [
            ("Nr wyceny:", self.report.quotation_id or "DRAFT"),
            ("Data:", self._format_date(self.report.quotation_date)),
            ("Ważna do:", self._format_date(self.report.valid_until) or "N/A"),
            ("Algorytm:", self.report.algorithm),
        ]
        
        for label, value in metadata:
            ws.cell(row=row, column=1, value=label).font = self.subheader_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Dane klienta
        row += 1
        ws.cell(row=row, column=1, value="DANE KLIENTA").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        customer_data = [
            ("Firma:", self.report.customer_company),
            ("Osoba:", self.report.customer_name),
            ("Email:", self.report.customer_email),
            ("Telefon:", self.report.customer_phone),
            ("NIP:", self.report.customer_nip),
        ]
        
        for label, value in customer_data:
            if value:
                ws.cell(row=row, column=1, value=label).font = self.subheader_font
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        # Statystyki
        row += 1
        ws.cell(row=row, column=1, value="STATYSTYKI").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        stats = [
            ("Liczba pozycji:", len(self.report.parts)),
            ("Łączna ilość detali:", self.report.total_parts_count),
            ("Materiały:", ", ".join(self.report.unique_materials)),
        ]
        
        if self.report.nesting:
            stats.extend([
                ("Arkusze:", self.report.nesting.total_sheets),
                ("Wykorzystanie:", f"{self.report.nesting.average_utilization * 100:.1f}%"),
            ])
        
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = self.subheader_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Koszty
        row += 1
        ws.cell(row=row, column=1, value="KOSZTY").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        costs = self.report.costs
        cost_items = [
            ("Materiał:", costs.material_cost),
            ("Cięcie:", costs.cutting_cost),
            ("Gięcie:", costs.bending_cost),
            ("Setup:", costs.setup_cost),
            ("Programowanie:", costs.programming_cost),
            ("Suma netto:", costs.subtotal),
            (f"Marża ({costs.margin_percent*100:.0f}%):", costs.margin_value),
        ]
        
        for label, value in cost_items:
            ws.cell(row=row, column=1, value=label).font = self.subheader_font
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = '#,##0.00 "PLN"'
            row += 1
        
        # Suma końcowa
        row += 1
        ws.cell(row=row, column=1, value="RAZEM:").font = self.total_font
        cell = ws.cell(row=row, column=2, value=costs.total)
        cell.font = self.total_font
        cell.number_format = '#,##0.00 "PLN"'
        cell.fill = self.total_fill
        
        # Szerokości kolumn
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _create_parts_sheet(self):
        """Arkusz detali"""
        ws = self.wb.create_sheet("Detale")
        
        # Nagłówki
        headers = ['Lp.', 'Nazwa', 'Materiał', 'Grubość [mm]', 'Szerokość [mm]', 
                   'Wysokość [mm]', 'Pole [mm²]', 'Obwód [mm]', 'Ilość', 
                   'Gięcie', 'Cena jedn.', 'Wartość', 'Plik 2D']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Dane
        for row_idx, part in enumerate(self.report.parts, 2):
            data = [
                row_idx - 1,
                part.name,
                part.material,
                part.thickness_mm,
                part.width_mm,
                part.height_mm,
                part.area_mm2,
                part.perimeter_mm,
                part.quantity,
                "Tak" if part.has_bending else "Nie",
                part.unit_cost,
                part.total_cost,
                Path(part.file_2d).name if part.file_2d else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                cell.font = self.normal_font
                
                # Formatowanie
                if col in [4, 5, 6]:  # Wymiary
                    cell.number_format = '0.00'
                    cell.alignment = self.right_align
                elif col in [7, 8]:  # Pole, obwód
                    cell.number_format = '#,##0.00'
                    cell.alignment = self.right_align
                elif col in [11, 12]:  # Ceny
                    cell.number_format = '#,##0.00 "PLN"'
                    cell.alignment = self.right_align
                elif col in [1, 3, 4, 9, 10]:  # Centrowane
                    cell.alignment = self.center_align
            
            # Naprzemienne kolory wierszy
            if row_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color='FFF9FAFB', end_color='FFF9FAFB', fill_type='solid'
                    )
        
        # Suma
        sum_row = len(self.report.parts) + 2
        ws.cell(row=sum_row, column=1, value="SUMA:").font = self.total_font
        
        sum_qty = ws.cell(row=sum_row, column=9, 
                          value=f"=SUM(I2:I{sum_row-1})")
        sum_qty.font = self.total_font
        
        sum_value = ws.cell(row=sum_row, column=12, 
                            value=f"=SUM(L2:L{sum_row-1})")
        sum_value.font = self.total_font
        sum_value.number_format = '#,##0.00 "PLN"'
        
        # Szerokości kolumn
        widths = [6, 35, 15, 12, 12, 12, 15, 12, 8, 8, 15, 15, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Autofiltr
        ws.auto_filter.ref = f"A1:M{len(self.report.parts) + 1}"
    
    def _create_nesting_sheet(self):
        """Arkusz nestingu"""
        ws = self.wb.create_sheet("Nesting")
        
        if not self.report.nesting:
            ws['A1'] = "Brak danych nestingu"
            return
        
        nesting = self.report.nesting
        
        # Statystyki główne
        ws.merge_cells('A1:D1')
        ws['A1'] = "WYNIKI NESTINGU"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color=self.COLOR_PRIMARY[2:])
        
        # Parametry
        row = 3
        params = [
            ("Algorytm:", self.report.algorithm),
            ("Kerf:", f"{self.report.kerf_width} mm"),
            ("Odstęp:", f"{self.report.part_spacing} mm"),
            ("Margines:", f"{self.report.sheet_margin} mm"),
        ]
        
        for label, value in params:
            ws.cell(row=row, column=1, value=label).font = self.subheader_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Podsumowanie
        row += 1
        ws.cell(row=row, column=1, value="PODSUMOWANIE").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        summary = [
            ("Liczba arkuszy:", nesting.total_sheets),
            ("Umieszczone detale:", nesting.total_parts),
            ("Średnie wykorzystanie:", f"{nesting.average_utilization * 100:.1f}%"),
            ("Całkowita powierzchnia:", f"{nesting.total_sheet_area_mm2 / 1_000_000:.3f} m²"),
            ("Wykorzystana powierzchnia:", f"{nesting.total_used_area_mm2 / 1_000_000:.3f} m²"),
            ("Odpad:", f"{nesting.total_waste_area_mm2 / 1_000_000:.3f} m²"),
        ]
        
        for label, value in summary:
            ws.cell(row=row, column=1, value=label).font = self.subheader_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Szczegóły arkuszy
        row += 2
        ws.cell(row=row, column=1, value="SZCZEGÓŁY ARKUSZY").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Nagłówki tabeli
        sheet_headers = ['Nr', 'Format', 'Materiał', 'Grubość', 'Detali', 'Wykorzystanie']
        for col, header in enumerate(sheet_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        row += 1
        
        # Dane arkuszy
        for sheet in nesting.sheets:
            data = [
                sheet.index,
                sheet.format_name,
                sheet.material,
                f"{sheet.thickness_mm}mm",
                sheet.parts_count,
                f"{sheet.utilization * 100:.1f}%"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.thin_border
                cell.alignment = self.center_align
            row += 1
        
        # Wykres wykorzystania
        if len(nesting.sheets) > 1:
            self._add_utilization_chart(ws, nesting, row + 2)
        
        # Szerokości
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
    
    def _add_utilization_chart(self, ws, nesting: NestingReport, start_row: int):
        """Dodaj wykres wykorzystania"""
        # Przygotuj dane do wykresu
        ws.cell(row=start_row, column=1, value="Arkusz")
        ws.cell(row=start_row, column=2, value="Wykorzystanie %")
        
        for i, sheet in enumerate(nesting.sheets, 1):
            ws.cell(row=start_row + i, column=1, value=f"Ark. {sheet.index}")
            ws.cell(row=start_row + i, column=2, value=sheet.utilization * 100)
        
        # Utwórz wykres
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Wykorzystanie arkuszy"
        chart.y_axis.title = "Wykorzystanie [%]"
        chart.x_axis.title = "Arkusz"
        
        data = Reference(ws, min_col=2, min_row=start_row, 
                        max_row=start_row + len(nesting.sheets))
        cats = Reference(ws, min_col=1, min_row=start_row + 1, 
                        max_row=start_row + len(nesting.sheets))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        ws.add_chart(chart, f"H{start_row}")
    
    def _create_costs_sheet(self):
        """Arkusz kosztów"""
        ws = self.wb.create_sheet("Koszty")
        
        costs = self.report.costs
        
        # Tytuł
        ws.merge_cells('A1:C1')
        ws['A1'] = "KALKULACJA KOSZTÓW"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color=self.COLOR_PRIMARY[2:])
        
        # Tabela kosztów
        row = 3
        cost_items = [
            ("Materiał", costs.material_cost, "Koszt arkuszy blach"),
            ("Cięcie laserowe", costs.cutting_cost, "Koszt pracy lasera"),
            ("Gięcie", costs.bending_cost, "Koszt operacji gięcia"),
            ("Przygotowanie (setup)", costs.setup_cost, "Ustawienie maszyny"),
            ("Programowanie", costs.programming_cost, "Przygotowanie programu CNC"),
        ]
        
        if costs.other_cost > 0:
            cost_items.append(("Inne", costs.other_cost, "Dodatkowe koszty"))
        
        # Nagłówki
        headers = ['Pozycja', 'Wartość', 'Opis']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        row += 1
        
        # Dane
        for name, value, desc in cost_items:
            ws.cell(row=row, column=1, value=name).border = self.thin_border
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = '#,##0.00 "PLN"'
            cell.alignment = self.right_align
            cell.border = self.thin_border
            ws.cell(row=row, column=3, value=desc).border = self.thin_border
            row += 1
        
        # Suma netto
        row += 1
        ws.cell(row=row, column=1, value="SUMA NETTO").font = self.subheader_font
        cell = ws.cell(row=row, column=2, value=costs.subtotal)
        cell.font = self.subheader_font
        cell.number_format = '#,##0.00 "PLN"'
        cell.alignment = self.right_align
        row += 1
        
        # Marża
        ws.cell(row=row, column=1, value=f"Marża ({costs.margin_percent*100:.0f}%)")
        cell = ws.cell(row=row, column=2, value=costs.margin_value)
        cell.number_format = '#,##0.00 "PLN"'
        cell.alignment = self.right_align
        row += 2
        
        # Suma końcowa
        ws.cell(row=row, column=1, value="RAZEM DO ZAPŁATY").font = self.total_font
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color=self.COLOR_SUCCESS, end_color=self.COLOR_SUCCESS, fill_type='solid'
        )
        ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        
        cell = ws.cell(row=row, column=2, value=costs.total)
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.number_format = '#,##0.00 "PLN"'
        cell.alignment = self.right_align
        cell.fill = PatternFill(start_color=self.COLOR_SUCCESS, 
                                end_color=self.COLOR_SUCCESS, fill_type='solid')
        
        # Wykres kołowy
        self._add_cost_pie_chart(ws, costs, row + 3)
        
        # Szerokości
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
    
    def _add_cost_pie_chart(self, ws, costs, start_row: int):
        """Dodaj wykres kołowy kosztów"""
        # Dane do wykresu
        chart_data = [
            ("Materiał", costs.material_cost),
            ("Cięcie", costs.cutting_cost),
            ("Gięcie", costs.bending_cost),
            ("Setup", costs.setup_cost),
            ("Programowanie", costs.programming_cost),
        ]
        
        # Filtruj zerowe
        chart_data = [(name, val) for name, val in chart_data if val > 0]
        
        if not chart_data:
            return
        
        # Zapisz dane
        ws.cell(row=start_row, column=1, value="Kategoria")
        ws.cell(row=start_row, column=2, value="Wartość")
        
        for i, (name, value) in enumerate(chart_data, 1):
            ws.cell(row=start_row + i, column=1, value=name)
            ws.cell(row=start_row + i, column=2, value=value)
        
        # Utwórz wykres
        chart = PieChart()
        chart.title = "Struktura kosztów"
        
        data = Reference(ws, min_col=2, min_row=start_row, 
                        max_row=start_row + len(chart_data))
        labels = Reference(ws, min_col=1, min_row=start_row + 1, 
                          max_row=start_row + len(chart_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        ws.add_chart(chart, "E3")


def generate_excel_report(report: QuotationReport, output_path: str) -> bool:
    """Funkcja pomocnicza do generowania Excel"""
    generator = ExcelReportGenerator(report)
    return generator.generate(output_path)
