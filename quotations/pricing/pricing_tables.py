"""
NewERP - Pricing Tables
=======================
Cenniki materiałów i kosztów cięcia laserowego.

Obsługuje:
- Import z plików XLSX
- Domyślne cenniki w kodzie
- Wyszukiwanie cen po materiale i grubości
"""

import os
import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from enum import Enum

logger = logging.getLogger(__name__)


@dataclass
class MaterialPrice:
    """Cena materiału"""
    material: str           # Nazwa/gatunek np. "1.4301", "S235"
    thickness_mm: float     # Grubość
    price_per_kg: float     # Cena za kg [PLN]
    price_per_m2: float = 0.0   # Cena za m² (opcjonalnie)
    density_kg_m3: float = 7850  # Gęstość [kg/m³]
    
    @property
    def price_per_m2_calculated(self) -> float:
        """Oblicz cenę za m² na podstawie grubości i gęstości"""
        if self.price_per_m2 > 0:
            return self.price_per_m2
        # kg/m² = grubość[m] × gęstość[kg/m³]
        kg_per_m2 = (self.thickness_mm / 1000) * self.density_kg_m3
        return kg_per_m2 * self.price_per_kg


@dataclass
class CuttingRate:
    """Stawka cięcia laserowego"""
    material_type: str      # Typ: "steel", "stainless", "aluminum"
    thickness_mm: float     # Grubość
    speed_mm_min: float     # Prędkość cięcia [mm/min]
    cost_per_meter: float   # Koszt za metr cięcia [PLN/m]
    cost_per_hour: float = 0.0   # Koszt za godzinę (opcjonalnie)
    piercing_time_s: float = 0.5  # Czas przebicia [s]


@dataclass 
class BendingRate:
    """Stawka gięcia"""
    thickness_mm: float     # Grubość
    cost_per_bend: float    # Koszt za gięcie [PLN]
    min_length_mm: float = 0.0   # Min długość gięcia
    max_length_mm: float = 3000.0  # Max długość gięcia


class PricingTables:
    """
    Zarządzanie cennikami.
    
    Użycie:
        tables = PricingTables()
        tables.load_from_xlsx("cenniki.xlsx")
        
        price = tables.get_material_price("1.4301", 2.0)
        cutting = tables.get_cutting_rate("stainless", 2.0)
    """
    
    def __init__(self):
        self.material_prices: Dict[Tuple[str, float], MaterialPrice] = {}
        self.cutting_rates: Dict[Tuple[str, float], CuttingRate] = {}
        self.bending_rates: Dict[float, BendingRate] = {}
        
        # Stawki podstawowe
        self.laser_hourly_rate: float = 350.0  # PLN/h
        self.bending_hourly_rate: float = 200.0  # PLN/h
        self.setup_cost: float = 50.0  # PLN
        self.programming_cost: float = 10.0  # PLN/detal
        
        # Załaduj domyślne cenniki
        self._load_defaults()
    
    def _load_defaults(self):
        """Załaduj domyślne cenniki"""
        # === MATERIAŁY ===
        
        # Stal czarna S235
        for thickness, price in [
            (0.5, 4.20), (0.8, 4.30), (1.0, 4.40), (1.5, 4.50),
            (2.0, 4.50), (2.5, 4.60), (3.0, 4.70), (4.0, 4.80),
            (5.0, 4.90), (6.0, 5.00), (8.0, 5.10), (10.0, 5.20),
            (12.0, 5.30), (15.0, 5.50), (20.0, 5.80), (25.0, 6.00)
        ]:
            self.material_prices[("S235", thickness)] = MaterialPrice(
                material="S235", thickness_mm=thickness, 
                price_per_kg=price, density_kg_m3=7850
            )
        
        # Stal S355
        for thickness, price in [
            (2.0, 5.00), (3.0, 5.10), (4.0, 5.20), (5.0, 5.30),
            (6.0, 5.40), (8.0, 5.50), (10.0, 5.60), (12.0, 5.80),
            (15.0, 6.00), (20.0, 6.50)
        ]:
            self.material_prices[("S355", thickness)] = MaterialPrice(
                material="S355", thickness_mm=thickness,
                price_per_kg=price, density_kg_m3=7850
            )
        
        # DC01
        for thickness, price in [
            (0.5, 5.00), (0.8, 5.10), (1.0, 5.20), (1.5, 5.30),
            (2.0, 5.40), (2.5, 5.50), (3.0, 5.60)
        ]:
            self.material_prices[("DC01", thickness)] = MaterialPrice(
                material="DC01", thickness_mm=thickness,
                price_per_kg=price, density_kg_m3=7850
            )
        
        # Stal nierdzewna 1.4301 (INOX 304)
        for thickness, price in [
            (0.5, 16.00), (0.8, 16.50), (1.0, 17.00), (1.5, 17.50),
            (2.0, 18.00), (2.5, 18.50), (3.0, 19.00), (4.0, 20.00),
            (5.0, 21.00), (6.0, 22.00), (8.0, 24.00), (10.0, 26.00),
            (12.0, 28.00), (15.0, 32.00)
        ]:
            self.material_prices[("1.4301", thickness)] = MaterialPrice(
                material="1.4301", thickness_mm=thickness,
                price_per_kg=price, density_kg_m3=7900
            )
            # Alias INOX304
            self.material_prices[("INOX304", thickness)] = MaterialPrice(
                material="INOX304", thickness_mm=thickness,
                price_per_kg=price, density_kg_m3=7900
            )
        
        # 1.4404 (316L)
        for thickness, price in [
            (1.0, 22.00), (1.5, 22.50), (2.0, 23.00), (3.0, 24.00),
            (4.0, 25.00), (5.0, 26.00), (6.0, 28.00), (8.0, 30.00),
            (10.0, 32.00)
        ]:
            self.material_prices[("1.4404", thickness)] = MaterialPrice(
                material="1.4404", thickness_mm=thickness,
                price_per_kg=price, density_kg_m3=7950
            )
        
        # Aluminium
        for thickness, price in [
            (1.0, 22.00), (1.5, 23.00), (2.0, 24.00), (3.0, 25.00),
            (4.0, 26.00), (5.0, 27.00), (6.0, 28.00), (8.0, 30.00),
            (10.0, 32.00)
        ]:
            self.material_prices[("ALU", thickness)] = MaterialPrice(
                material="ALU", thickness_mm=thickness,
                price_per_kg=price, density_kg_m3=2700
            )
        
        # === CIĘCIE LASEROWE ===
        
        # Stal czarna
        steel_cutting = [
            (0.5, 12000, 0.15), (0.8, 10000, 0.18), (1.0, 9000, 0.20),
            (1.5, 7500, 0.25), (2.0, 6000, 0.30), (2.5, 5000, 0.35),
            (3.0, 4200, 0.42), (4.0, 3200, 0.55), (5.0, 2500, 0.70),
            (6.0, 2000, 0.90), (8.0, 1400, 1.30), (10.0, 1000, 1.80),
            (12.0, 750, 2.40), (15.0, 500, 3.50), (20.0, 300, 5.50),
            (25.0, 180, 8.00)
        ]
        for thickness, speed, cost in steel_cutting:
            self.cutting_rates[("steel", thickness)] = CuttingRate(
                material_type="steel", thickness_mm=thickness,
                speed_mm_min=speed, cost_per_meter=cost,
                piercing_time_s=0.3 + thickness * 0.1
            )
        
        # Stal nierdzewna (wolniej, drożej)
        for thickness, speed, cost in steel_cutting:
            self.cutting_rates[("stainless", thickness)] = CuttingRate(
                material_type="stainless", thickness_mm=thickness,
                speed_mm_min=speed * 0.7,  # 30% wolniej
                cost_per_meter=cost * 1.4,  # 40% drożej
                piercing_time_s=0.5 + thickness * 0.15
            )
        
        # Aluminium (szybciej)
        for thickness, speed, cost in steel_cutting:
            if thickness <= 10:
                self.cutting_rates[("aluminum", thickness)] = CuttingRate(
                    material_type="aluminum", thickness_mm=thickness,
                    speed_mm_min=speed * 1.3,  # 30% szybciej
                    cost_per_meter=cost * 1.1,  # 10% drożej
                    piercing_time_s=0.2 + thickness * 0.05
                )
        
        # === GIĘCIE ===
        for thickness, cost in [
            (0.5, 1.50), (0.8, 1.80), (1.0, 2.00), (1.5, 2.30),
            (2.0, 2.50), (2.5, 2.80), (3.0, 3.00), (4.0, 3.50),
            (5.0, 4.00), (6.0, 4.50), (8.0, 5.50), (10.0, 7.00),
            (12.0, 9.00), (15.0, 12.00)
        ]:
            self.bending_rates[thickness] = BendingRate(
                thickness_mm=thickness, cost_per_bend=cost
            )
    
    def load_from_xlsx(self, filepath: str | Path) -> bool:
        """
        Załaduj cenniki z pliku XLSX.
        
        Oczekiwana struktura:
        - Arkusz "Materiały": kolumny Material, Grubość, Cena/kg, Gęstość
        - Arkusz "Cięcie": kolumny Typ, Grubość, Prędkość, Cena/m
        - Arkusz "Gięcie": kolumny Grubość, Cena/gięcie
        - Arkusz "Stawki": Laser/h, Gięcie/h, Setup, Programowanie
        """
        try:
            import openpyxl
            
            filepath = Path(filepath)
            if not filepath.exists():
                logger.error(f"File not found: {filepath}")
                return False
            
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            # === Materiały ===
            if "Materiały" in wb.sheetnames or "Materialy" in wb.sheetnames:
                sheet_name = "Materiały" if "Materiały" in wb.sheetnames else "Materialy"
                ws = wb[sheet_name]
                
                # Znajdź nagłówki
                headers = {}
                for col, cell in enumerate(ws[1], 1):
                    if cell.value:
                        headers[cell.value.lower().strip()] = col
                
                # Wymagane kolumny
                mat_col = headers.get('materiał') or headers.get('material') or headers.get('gatunek')
                thick_col = headers.get('grubość') or headers.get('grubosc') or headers.get('thickness')
                price_col = headers.get('cena/kg') or headers.get('cena') or headers.get('price')
                
                if mat_col and thick_col and price_col:
                    for row in ws.iter_rows(min_row=2):
                        material = row[mat_col - 1].value
                        thickness = row[thick_col - 1].value
                        price = row[price_col - 1].value
                        
                        if material and thickness and price:
                            try:
                                thickness = float(str(thickness).replace(',', '.'))
                                price = float(str(price).replace(',', '.'))
                                
                                # Gęstość (opcjonalna)
                                density = 7850
                                density_col = headers.get('gęstość') or headers.get('gestosc') or headers.get('density')
                                if density_col and row[density_col - 1].value:
                                    density = float(str(row[density_col - 1].value).replace(',', '.'))
                                
                                material = str(material).strip().upper()
                                self.material_prices[(material, thickness)] = MaterialPrice(
                                    material=material,
                                    thickness_mm=thickness,
                                    price_per_kg=price,
                                    density_kg_m3=density
                                )
                                logger.debug(f"Loaded material: {material} {thickness}mm @ {price} PLN/kg")
                            except (ValueError, TypeError) as e:
                                logger.warning(f"Invalid row in Materiały: {e}")
            
            # === Cięcie ===
            if "Cięcie" in wb.sheetnames or "Ciecie" in wb.sheetnames:
                sheet_name = "Cięcie" if "Cięcie" in wb.sheetnames else "Ciecie"
                ws = wb[sheet_name]
                
                headers = {}
                for col, cell in enumerate(ws[1], 1):
                    if cell.value:
                        headers[cell.value.lower().strip()] = col
                
                type_col = headers.get('typ') or headers.get('type') or headers.get('materiał')
                thick_col = headers.get('grubość') or headers.get('grubosc') or headers.get('thickness')
                speed_col = headers.get('prędkość') or headers.get('predkosc') or headers.get('speed')
                cost_col = headers.get('cena/m') or headers.get('cena') or headers.get('cost')
                
                if thick_col and (speed_col or cost_col):
                    for row in ws.iter_rows(min_row=2):
                        mat_type = row[type_col - 1].value if type_col else "steel"
                        thickness = row[thick_col - 1].value
                        speed = row[speed_col - 1].value if speed_col else 5000
                        cost = row[cost_col - 1].value if cost_col else 0.5
                        
                        if thickness:
                            try:
                                mat_type = str(mat_type).lower().strip() if mat_type else "steel"
                                thickness = float(str(thickness).replace(',', '.'))
                                speed = float(str(speed).replace(',', '.')) if speed else 5000
                                cost = float(str(cost).replace(',', '.')) if cost else 0.5
                                
                                self.cutting_rates[(mat_type, thickness)] = CuttingRate(
                                    material_type=mat_type,
                                    thickness_mm=thickness,
                                    speed_mm_min=speed,
                                    cost_per_meter=cost
                                )
                                logger.debug(f"Loaded cutting: {mat_type} {thickness}mm @ {cost} PLN/m")
                            except (ValueError, TypeError) as e:
                                logger.warning(f"Invalid row in Cięcie: {e}")
            
            # === Gięcie ===
            if "Gięcie" in wb.sheetnames or "Giecie" in wb.sheetnames:
                sheet_name = "Gięcie" if "Gięcie" in wb.sheetnames else "Giecie"
                ws = wb[sheet_name]
                
                headers = {}
                for col, cell in enumerate(ws[1], 1):
                    if cell.value:
                        headers[cell.value.lower().strip()] = col
                
                thick_col = headers.get('grubość') or headers.get('grubosc') or headers.get('thickness')
                cost_col = headers.get('cena') or headers.get('cena/gięcie') or headers.get('cost')
                
                if thick_col and cost_col:
                    for row in ws.iter_rows(min_row=2):
                        thickness = row[thick_col - 1].value
                        cost = row[cost_col - 1].value
                        
                        if thickness and cost:
                            try:
                                thickness = float(str(thickness).replace(',', '.'))
                                cost = float(str(cost).replace(',', '.'))
                                
                                self.bending_rates[thickness] = BendingRate(
                                    thickness_mm=thickness,
                                    cost_per_bend=cost
                                )
                                logger.debug(f"Loaded bending: {thickness}mm @ {cost} PLN")
                            except (ValueError, TypeError) as e:
                                logger.warning(f"Invalid row in Gięcie: {e}")
            
            # === Stawki ===
            if "Stawki" in wb.sheetnames:
                ws = wb["Stawki"]
                for row in ws.iter_rows(min_row=1, max_col=2):
                    if row[0].value and row[1].value:
                        name = str(row[0].value).lower().strip()
                        value = float(str(row[1].value).replace(',', '.'))
                        
                        if 'laser' in name:
                            self.laser_hourly_rate = value
                        elif 'gięcie' in name or 'giecie' in name or 'bending' in name:
                            self.bending_hourly_rate = value
                        elif 'setup' in name:
                            self.setup_cost = value
                        elif 'programowanie' in name or 'programming' in name:
                            self.programming_cost = value
            
            wb.close()
            logger.info(f"Loaded pricing from {filepath}")
            return True
            
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Error loading XLSX: {e}")
            return False
    
    def load_materials_xlsx(self, filepath: str | Path) -> bool:
        """
        Załaduj cennik materiałów z pliku XLSX (format AVE).
        
        Struktura:
        format | material | thickness | price | date | from | note
        """
        try:
            import openpyxl
            
            filepath = Path(filepath)
            if not filepath.exists():
                logger.error(f"File not found: {filepath}")
                return False
            
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Znajdź nagłówki
            headers = {}
            for col, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[str(cell.value).lower().strip()] = col
            
            mat_col = headers.get('material', 2)
            thick_col = headers.get('thickness', 3)
            price_col = headers.get('price', 4)
            
            # Gęstości materiałów
            densities = {
                'S235': 7850, 'S355': 7850, 'DC01': 7850,
                '1.4301': 7900, '1.4404': 7950, '1.4571': 7950,
                'ALUMINIUM': 2700, '42CRM04': 7850, '42CRMO4': 7850,
                'CORTEN': 7850, 'P265GH': 7850
            }
            
            count = 0
            for row in ws.iter_rows(min_row=2):
                material = row[mat_col - 1].value
                thickness = row[thick_col - 1].value
                price = row[price_col - 1].value
                
                if material and thickness and price:
                    try:
                        material = str(material).strip().upper()
                        thickness = float(str(thickness).replace(',', '.'))
                        price = float(str(price).replace(',', '.'))
                        density = densities.get(material, 7850)
                        
                        self.material_prices[(material, thickness)] = MaterialPrice(
                            material=material,
                            thickness_mm=thickness,
                            price_per_kg=price,
                            density_kg_m3=density
                        )
                        count += 1
                    except (ValueError, TypeError):
                        pass
            
            wb.close()
            logger.info(f"Loaded {count} material prices from {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error loading materials XLSX: {e}")
            return False
    
    def load_cutting_xlsx(self, filepath: str | Path) -> bool:
        """
        Załaduj cennik cięcia z pliku XLSX (format AVE).
        
        Struktura:
        thickness | material | gas | price | speed | hour_price | utilization
        """
        try:
            import openpyxl
            
            filepath = Path(filepath)
            if not filepath.exists():
                logger.error(f"File not found: {filepath}")
                return False
            
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Znajdź nagłówki
            headers = {}
            for col, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[str(cell.value).lower().strip()] = col
            
            thick_col = headers.get('thickness', 1)
            mat_col = headers.get('material', 2)
            price_col = headers.get('price', 4)
            speed_col = headers.get('speed', 5)
            hour_col = headers.get('hour_price', 6)
            
            count = 0
            for row in ws.iter_rows(min_row=2):
                thickness = row[thick_col - 1].value
                material = row[mat_col - 1].value
                price = row[price_col - 1].value
                speed = row[speed_col - 1].value
                hour_price = row[hour_col - 1].value if hour_col else 750
                
                if thickness and material and price:
                    try:
                        material = str(material).strip().upper()
                        thickness = float(str(thickness).replace(',', '.'))
                        price = float(str(price).replace(',', '.'))
                        # Prędkość w m/min -> mm/min
                        speed = float(str(speed).replace(',', '.')) * 1000 if speed else 5000
                        hour_price = float(str(hour_price).replace(',', '.')) if hour_price else 750
                        
                        self.cutting_rates[(material, thickness)] = CuttingRate(
                            material_type=material,
                            thickness_mm=thickness,
                            speed_mm_min=speed,
                            cost_per_meter=price,
                            cost_per_hour=hour_price
                        )
                        count += 1
                    except (ValueError, TypeError):
                        pass
            
            wb.close()
            logger.info(f"Loaded {count} cutting rates from {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error loading cutting XLSX: {e}")
            return False
    
    def get_material_price(self, material: str, thickness: float) -> Optional[MaterialPrice]:
        """Pobierz cenę materiału"""
        material = material.upper().strip()
        
        # Dokładne dopasowanie
        if (material, thickness) in self.material_prices:
            return self.material_prices[(material, thickness)]
        
        # Aliasy materiałów
        aliases = {
            'INOX': '1.4301', 'INOX304': '1.4301', '304': '1.4301',
            'INOX316': '1.4404', '316': '1.4404', '316L': '1.4404',
            'FE': 'S235', 'ST37': 'S235',
            'ALUMINIUM': 'ALU', 'AL': 'ALU'
        }
        
        if material in aliases:
            alt_material = aliases[material]
            if (alt_material, thickness) in self.material_prices:
                return self.material_prices[(alt_material, thickness)]
        
        # Znajdź najbliższą grubość
        same_material = [(m, t) for (m, t) in self.material_prices.keys() if m == material]
        if same_material:
            closest = min(same_material, key=lambda x: abs(x[1] - thickness))
            return self.material_prices[closest]
        
        return None
    
    def get_cutting_rate(self, material_type: str, thickness: float) -> Optional[CuttingRate]:
        """Pobierz stawkę cięcia"""
        material_type = material_type.lower().strip()
        
        # Dokładne dopasowanie
        if (material_type, thickness) in self.cutting_rates:
            return self.cutting_rates[(material_type, thickness)]
        
        # Mapowanie typów
        type_map = {
            '1.4301': 'stainless', '1.4404': 'stainless', 'inox': 'stainless', 'inox304': 'stainless',
            's235': 'steel', 's355': 'steel', 'dc01': 'steel', 'fe': 'steel',
            'alu': 'aluminum', 'aluminium': 'aluminum', 'al': 'aluminum'
        }
        
        if material_type in type_map:
            material_type = type_map[material_type]
        
        if (material_type, thickness) in self.cutting_rates:
            return self.cutting_rates[(material_type, thickness)]
        
        # Najbliższa grubość
        same_type = [(t, th) for (t, th) in self.cutting_rates.keys() if t == material_type]
        if same_type:
            closest = min(same_type, key=lambda x: abs(x[1] - thickness))
            return self.cutting_rates[closest]
        
        return None
    
    def get_bending_rate(self, thickness: float) -> Optional[BendingRate]:
        """Pobierz stawkę gięcia"""
        if thickness in self.bending_rates:
            return self.bending_rates[thickness]
        
        # Najbliższa grubość
        if self.bending_rates:
            closest = min(self.bending_rates.keys(), key=lambda t: abs(t - thickness))
            return self.bending_rates[closest]
        
        return None
    
    def list_materials(self) -> List[str]:
        """Lista dostępnych materiałów"""
        return sorted(set(m for m, t in self.material_prices.keys()))
    
    def list_thicknesses(self, material: str = None) -> List[float]:
        """Lista dostępnych grubości"""
        if material:
            return sorted([t for m, t in self.material_prices.keys() if m == material.upper()])
        return sorted(set(t for m, t in self.material_prices.keys()))


# Globalny singleton
_pricing_tables: Optional[PricingTables] = None


def get_pricing_tables() -> PricingTables:
    """Pobierz globalną instancję cenników"""
    global _pricing_tables
    if _pricing_tables is None:
        _pricing_tables = PricingTables()
        
        # Spróbuj załadować cenniki z data/pricing
        try:
            data_dir = Path(__file__).parent.parent.parent / "data" / "pricing"
            
            materials_file = data_dir / "materials_prices.xlsx"
            if materials_file.exists():
                _pricing_tables.load_materials_xlsx(materials_file)
                logger.info(f"Auto-loaded materials from {materials_file}")
            
            cutting_file = data_dir / "cutting_prices.xlsx"
            if cutting_file.exists():
                _pricing_tables.load_cutting_xlsx(cutting_file)
                logger.info(f"Auto-loaded cutting rates from {cutting_file}")
                
        except Exception as e:
            logger.warning(f"Could not auto-load pricing files: {e}")
    
    return _pricing_tables


def load_pricing_from_xlsx(filepath: str | Path) -> bool:
    """Załaduj cenniki z XLSX"""
    tables = get_pricing_tables()
    return tables.load_from_xlsx(filepath)


# ============================================================
# CLI Test
# ============================================================

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    
    tables = PricingTables()
    
    print("="*60)
    print("DOSTĘPNE MATERIAŁY")
    print("="*60)
    
    for material in tables.list_materials():
        thicknesses = tables.list_thicknesses(material)
        print(f"\n{material}:")
        for t in thicknesses[:5]:
            price = tables.get_material_price(material, t)
            if price:
                print(f"  {t}mm: {price.price_per_kg:.2f} PLN/kg = {price.price_per_m2_calculated:.2f} PLN/m²")
    
    print("\n" + "="*60)
    print("STAWKI CIĘCIA (stal)")
    print("="*60)
    
    for (mat_type, thickness), rate in sorted(tables.cutting_rates.items()):
        if mat_type == "steel" and thickness <= 10:
            print(f"  {thickness}mm: {rate.speed_mm_min:.0f} mm/min, {rate.cost_per_meter:.2f} PLN/m")
