"""
Excel Price Importer
====================
Import cenników z plików XLSX do bazy danych.
"""

import os
import logging
from typing import List, Dict, Tuple, Optional
from datetime import datetime, date
from dataclasses import dataclass

logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed. Run: pip install openpyxl")


@dataclass
class ImportResult:
    """Wynik importu"""
    success: bool
    imported: int = 0
    updated: int = 0
    failed: int = 0
    errors: List[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []


class ExcelPriceImporter:
    """Importer cenników z plików Excel"""
    
    # Mapowanie nagłówków Excel na pola bazy
    MATERIAL_HEADERS = {
        'format': ['format', 'arkusz', 'sheet', 'wymiar'],
        'material': ['material', 'materiał', 'mat', 'gatunek'],
        'thickness': ['thickness', 'grubość', 'grubosc', 'gr', 'mm', 'grubość [mm]'],
        'price_per_kg': ['price', 'cena', 'price_per_kg', 'pln/kg', 'pln', 'cena [pln/kg]'],
        'source': ['source', 'from', 'źródło', 'dostawca', 'supplier'],
        'note': ['note', 'notes', 'uwagi', 'uwaga', 'komentarz'],
        'valid_from': ['date', 'data', 'valid_from', 'od', 'from_date', 'data od'],
    }
    
    CUTTING_HEADERS = {
        'material': ['material', 'materiał', 'mat'],
        'thickness': ['thickness', 'grubość', 'grubosc', 'gr', 'mm', 'grubość [mm]'],
        'gas': ['gas', 'gaz'],
        'cutting_speed': ['speed', 'prędkość', 'predkosc', 'cutting_speed', 'm/min', 'prędkość [m/min]'],
        'hour_price': ['hour_price', 'cena_godz', 'hourly', 'pln/h', 'cena/h [pln]', 'cena/h'],
        'utilization': ['utilization', 'wykorzystanie', 'util', 'wyk'],
        'price_per_meter': ['price', 'cena', 'price_per_meter', 'pln/m', 'cena/m [pln]', 'cena/m'],
        'note': ['note', 'notes', 'uwagi'],
    }
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel import")
    
    def read_material_prices(self, filepath: str) -> Tuple[List[Dict], List[str]]:
        """
        Wczytaj ceny materiałów z pliku Excel.
        
        Returns:
            Tuple[records, errors]
        """
        records = []
        errors = []
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Znajdź mapowanie kolumn
            header_map = self._find_headers(ws, self.MATERIAL_HEADERS)
            
            if 'material' not in header_map or 'thickness' not in header_map:
                errors.append("Brak wymaganych kolumn: material, thickness")
                return records, errors
            
            # Przetwórz wiersze
            for row_idx in range(2, ws.max_row + 1):
                try:
                    record = self._parse_material_row(ws, row_idx, header_map)
                    if record:
                        records.append(record)
                except Exception as e:
                    errors.append(f"Wiersz {row_idx}: {e}")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Błąd odczytu pliku: {e}")
        
        logger.info(f"Read {len(records)} material price records, {len(errors)} errors")
        return records, errors
    
    def read_cutting_prices(self, filepath: str) -> Tuple[List[Dict], List[str]]:
        """
        Wczytaj ceny cięcia z pliku Excel.
        
        Returns:
            Tuple[records, errors]
        """
        records = []
        errors = []
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Znajdź mapowanie kolumn
            header_map = self._find_headers(ws, self.CUTTING_HEADERS)
            
            if 'material' not in header_map or 'thickness' not in header_map:
                errors.append("Brak wymaganych kolumn: material, thickness")
                return records, errors
            
            # Przetwórz wiersze
            for row_idx in range(2, ws.max_row + 1):
                try:
                    record = self._parse_cutting_row(ws, row_idx, header_map)
                    if record:
                        records.append(record)
                except Exception as e:
                    errors.append(f"Wiersz {row_idx}: {e}")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Błąd odczytu pliku: {e}")
        
        logger.info(f"Read {len(records)} cutting price records, {len(errors)} errors")
        return records, errors
    
    def _find_headers(self, ws, header_mapping: Dict) -> Dict[str, int]:
        """Znajdź mapowanie nagłówków na numery kolumn"""
        result = {}
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(1, col_idx).value
            if header is None:
                continue
            
            header_lower = str(header).lower().strip()
            
            for field_name, possible_names in header_mapping.items():
                if header_lower in possible_names:
                    result[field_name] = col_idx
                    break
        
        logger.debug(f"Header mapping: {result}")
        return result
    
    def _parse_material_row(self, ws, row_idx: int, header_map: Dict) -> Optional[Dict]:
        """Parsuj wiersz ceny materiału"""
        material = self._get_cell_value(ws, row_idx, header_map.get('material'))
        thickness = self._get_cell_value(ws, row_idx, header_map.get('thickness'))
        price = self._get_cell_value(ws, row_idx, header_map.get('price_per_kg'))
        
        if not material or thickness is None:
            return None
        
        # Konwertuj wartości
        try:
            thickness = float(thickness)
        except (ValueError, TypeError):
            return None
        
        try:
            price = float(price) if price else 0.0
        except (ValueError, TypeError):
            price = 0.0
        
        record = {
            'material': str(material).strip().upper(),
            'thickness': thickness,
            'price_per_kg': price,
            'format': str(self._get_cell_value(ws, row_idx, header_map.get('format')) or '1500x3000'),
            'source': self._get_cell_value(ws, row_idx, header_map.get('source')),
            'note': self._get_cell_value(ws, row_idx, header_map.get('note')),
        }
        
        # Data
        valid_from = self._get_cell_value(ws, row_idx, header_map.get('valid_from'))
        if valid_from:
            if isinstance(valid_from, datetime):
                record['valid_from'] = valid_from.date().isoformat()
            elif isinstance(valid_from, date):
                record['valid_from'] = valid_from.isoformat()
            else:
                record['valid_from'] = date.today().isoformat()
        else:
            record['valid_from'] = date.today().isoformat()
        
        return record
    
    def _parse_cutting_row(self, ws, row_idx: int, header_map: Dict) -> Optional[Dict]:
        """Parsuj wiersz ceny cięcia"""
        material = self._get_cell_value(ws, row_idx, header_map.get('material'))
        thickness = self._get_cell_value(ws, row_idx, header_map.get('thickness'))
        
        if not material or thickness is None:
            return None
        
        try:
            thickness = float(thickness)
        except (ValueError, TypeError):
            return None
        
        # Pobierz parametry
        speed = self._get_cell_value(ws, row_idx, header_map.get('cutting_speed'))
        hour_price = self._get_cell_value(ws, row_idx, header_map.get('hour_price'))
        utilization = self._get_cell_value(ws, row_idx, header_map.get('utilization'))
        price_per_meter = self._get_cell_value(ws, row_idx, header_map.get('price_per_meter'))
        gas = self._get_cell_value(ws, row_idx, header_map.get('gas'))
        
        record = {
            'material': str(material).strip().upper(),
            'thickness': thickness,
            'gas': str(gas).strip().upper() if gas else 'N',
            'cutting_speed': float(speed) if speed else None,
            'hour_price': float(hour_price) if hour_price else 750.0,
            'utilization': float(utilization) if utilization else 0.65,
            'note': self._get_cell_value(ws, row_idx, header_map.get('note')),
            'valid_from': date.today().isoformat(),
        }
        
        # Cena za metr - jeśli podana bezpośrednio
        if price_per_meter:
            try:
                # Sprawdź czy to formuła
                if isinstance(price_per_meter, str) and price_per_meter.startswith('='):
                    # Oblicz z formuły
                    if record['cutting_speed'] and record['utilization']:
                        record['price_per_meter'] = record['hour_price'] / (
                            record['cutting_speed'] * 60 * record['utilization']
                        )
                        record['price_manual'] = False
                else:
                    record['price_per_meter'] = float(price_per_meter)
                    record['price_manual'] = True
            except (ValueError, TypeError):
                pass
        
        return record
    
    def _get_cell_value(self, ws, row: int, col: Optional[int]):
        """Pobierz wartość komórki"""
        if col is None:
            return None
        return ws.cell(row, col).value
    
    def detect_file_type(self, filepath: str) -> Optional[str]:
        """
        Wykryj typ cennika na podstawie nagłówków.
        
        Returns:
            'materials', 'cutting', or None
        """
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Zbierz nagłówki
            headers = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(1, col_idx).value
                if val:
                    headers.append(str(val).lower().strip())
            
            wb.close()
            
            # Sprawdź charakterystyczne nagłówki
            if any(h in headers for h in ['gas', 'gaz', 'speed', 'prędkość', 'hour_price']):
                return 'cutting'
            
            if any(h in headers for h in ['format', 'arkusz', 'price_per_kg', 'pln/kg']):
                return 'materials'
            
            # Domyślnie - materiały jeśli jest cena
            if any(h in headers for h in ['price', 'cena']):
                return 'materials'
            
            return None
            
        except Exception as e:
            logger.error(f"Error detecting file type: {e}")
            return None


class PriceExporter:
    """Eksporter cenników do Excel"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export")
    
    def export_material_prices(self, records: List[Dict], filepath: str) -> bool:
        """Eksportuj ceny materiałów do Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ceny materiałów"
            
            # Nagłówki
            headers = ['format', 'material', 'thickness', 'price_per_kg', 'source', 'note', 'valid_from']
            for col, header in enumerate(headers, 1):
                ws.cell(1, col, header)
            
            # Dane
            for row_idx, record in enumerate(records, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row_idx, col_idx, record.get(header))
            
            # Formatowanie
            self._format_worksheet(ws)
            
            wb.save(filepath)
            wb.close()
            return True
            
        except Exception as e:
            logger.error(f"Export error: {e}")
            return False
    
    def export_cutting_prices(self, records: List[Dict], filepath: str) -> bool:
        """Eksportuj ceny cięcia do Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ceny cięcia"
            
            # Nagłówki
            headers = ['material', 'thickness', 'gas', 'cutting_speed', 'hour_price', 
                      'utilization', 'price_per_meter', 'note', 'valid_from']
            for col, header in enumerate(headers, 1):
                ws.cell(1, col, header)
            
            # Dane
            for row_idx, record in enumerate(records, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row_idx, col_idx, record.get(header))
            
            self._format_worksheet(ws)
            
            wb.save(filepath)
            wb.close()
            return True
            
        except Exception as e:
            logger.error(f"Export error: {e}")
            return False
    
    def _format_worksheet(self, ws):
        """Formatuj arkusz"""
        from openpyxl.styles import Font, PatternFill
        
        # Nagłówki - pogrubienie + tło
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-szerokość kolumn
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
